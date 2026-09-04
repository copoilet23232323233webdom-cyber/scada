import flet as ft
from datetime import datetime
import os, time, threading, socket, sqlite3, json
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4
import csv, io

# ═══════════════════════════════════════════════════════════════
#  CONFIGURACION
# ═══════════════════════════════════════════════════════════════
PUERTO          = 502
REG_STATUS      = 57624
REG_VOLTAJE     = 57625
REG_STRINGS     = 57627
NUM_STRINGS     = 32
TIMEOUT_MS      = 2000
REINTENTOS      = 3
DELAY_REINTENTO = 0.5
MAX_WORKERS     = 10
DB_PATH         = os.path.join(os.path.expanduser("~"), ".webdom_scanner.db")

CLAVES_IP = ["ipgateway","gatewayip","ipgw","gwip","ipdispositivo","ipaddress","ipred",
             "ip","direccion","address","host","red","network","gateway","gw"]
CLAVES_ID = ["modbusid","slaveid","idmodbus","idnodo","idslave","unitid","modbusslave",
             "modbusunit","numeronodo","numerodispositivo","idesclave","identificador",
             "id","slave","nodo","node","unit","dispositivo","device","modbus"]
CLAVES_CAJA = ["caja","armario","rack","box","cuadro","panel","cabinet","ubicacion",
               "location","numero","num","ref","referencia","tag","numcaja","numerocaja",
               "cajanum","cajano","ncaja","numarmario","armarionum","etiqueta","label",
               "nombre","equipo","descripcion"]
CLAVES_STRINGS = ["nrostring","numstring","nstring","strings","string","nstrings",
                  "numstrings","numerostring","numerostrings","cantidadstrings",
                  "cantidadstring","nrostrings","stringsporbox","stringsporcaja","totalstrings"]

# ═══════════════════════════════════════════════════════════════
#  BASE DE DATOS LOCAL
# ═══════════════════════════════════════════════════════════════
def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS plantas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        creada TEXT DEFAULT (datetime('now')),
        activa INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS gateways (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        planta_id INTEGER NOT NULL,
        ip TEXT NOT NULL,
        id_inicio INTEGER NOT NULL,
        id_fin INTEGER NOT NULL,
        FOREIGN KEY(planta_id) REFERENCES plantas(id)
    );
    CREATE TABLE IF NOT EXISTS cbt_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        planta_id INTEGER NOT NULL,
        gateway_ip TEXT NOT NULL,
        modbus_id INTEGER NOT NULL,
        mac TEXT NOT NULL,
        UNIQUE(planta_id, gateway_ip, modbus_id)
    );
    """)
    conn.commit(); conn.close()

def get_db():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def db_get_plantas():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id, nombre FROM plantas WHERE activa=1 ORDER BY nombre")
    rows = c.fetchall()
    result = []
    for pid, nombre in rows:
        c.execute("SELECT ip, id_inicio, id_fin FROM gateways WHERE planta_id=?", (pid,))
        gws = [{"ip": r[0], "id_inicio": r[1], "id_fin": r[2]} for r in c.fetchall()]
        result.append({"id": pid, "nombre": nombre, "gateways": gws})
    conn.close()
    return result

def db_guardar_planta(nombre, gateways):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id FROM plantas WHERE nombre=?", (nombre,))
    row = c.fetchone()
    if row:
        pid = row[0]
        c.execute("DELETE FROM gateways WHERE planta_id=?", (pid,))
    else:
        c.execute("INSERT INTO plantas(nombre) VALUES(?)", (nombre,))
        pid = c.lastrowid
    for gw in gateways:
        c.execute("INSERT INTO gateways(planta_id,ip,id_inicio,id_fin) VALUES(?,?,?,?)",
                  (pid, gw[0], gw[1], gw[2]))
    conn.commit(); conn.close()
    return pid

def db_eliminar_planta(pid):
    conn = get_db()
    conn.execute("UPDATE plantas SET activa=0 WHERE id=?", (pid,))
    conn.commit(); conn.close()

def db_guardar_cbt(planta_id, gateway_ip, macs_dict):
    """macs_dict: {modbus_id: mac_str}"""
    conn = get_db(); c = conn.cursor()
    for mid, mac in macs_dict.items():
        c.execute("INSERT OR REPLACE INTO cbt_entries(planta_id,gateway_ip,modbus_id,mac) VALUES(?,?,?,?)",
                  (planta_id, gateway_ip, mid, mac))
    conn.commit(); conn.close()

def db_get_cbt(planta_id):
    """Devuelve {(gateway_ip, modbus_id): mac}"""
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT gateway_ip, modbus_id, mac FROM cbt_entries WHERE planta_id=?", (planta_id,))
    rows = c.fetchall(); conn.close()
    return {(r[0], r[1]): r[2] for r in rows}

# ═══════════════════════════════════════════════════════════════
#  MODBUS TCP
# ═══════════════════════════════════════════════════════════════
def modbus_read(ip, port, unit_id, register, count=1, timeout=2):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(timeout)
    try:
        sock.connect((ip, port))
        pdu = bytes([0x00,0x01,0x00,0x00,0x00,0x06, unit_id&0xFF, 0x03,
                     (register>>8)&0xFF, register&0xFF, (count>>8)&0xFF, count&0xFF])
        sock.sendall(pdu)
        buf = b""
        deadline = time.time() + timeout
        while len(buf) < 9 and time.time() < deadline:
            chunk = sock.recv(512)
            if not chunk: raise ConnectionError("Cerrado")
            buf += chunk
        if len(buf) < 9: raise TimeoutError("Incompleto")
        fc = buf[7]
        if fc & 0x80: raise Exception(f"Modbus err FC={fc:02X}")
        bc = buf[8]; total = 9+bc
        while len(buf) < total and time.time() < deadline:
            chunk = sock.recv(512)
            if not chunk: raise ConnectionError("Cerrado")
            buf += chunk
        regs = []
        for i in range(bc//2):
            o = 9+i*2
            raw = (buf[o]<<8)|buf[o+1]
            if raw >= 0x8000: raw -= 0x10000
            regs.append(raw)
        return regs
    finally:
        sock.close()

def leer_voltaje(ip, uid, timeout=2):
    try:
        r = modbus_read(ip, PUERTO, uid, REG_VOLTAJE, 1, timeout)
        return float(r[0])
    except: return None

def analizar_diagnostico(status_raw):
    fallos = []
    if not ((status_raw>>5)&1): fallos.append("LoRa FAIL")
    if not ((status_raw>>9)&1): fallos.append("Mem FAIL")
    return fallos

def leer_strings_caja(ip, uid, timeout=2):
    try:
        r = modbus_read(ip, PUERTO, uid, REG_STRINGS, NUM_STRINGS, timeout)
        return [x/10.0 for x in r]
    except: return None

def analizar_anomalias(corrientes, umbral_pct=30, num_strings=None):
    if corrientes is None: return []
    n = min(int(num_strings), NUM_STRINGS) if num_strings else NUM_STRINGS
    canales = list(range(n))
    activos = [corrientes[i] for i in canales if corrientes[i] > 0.5]
    if len(activos) < 2: return []
    media = sum(activos)/len(activos)
    if media < 1.0: return []
    umbral_v = media*(umbral_pct/100.0)
    res = []
    for i in canales:
        c = corrientes[i]
        if c <= 0.5:
            res.append({"string":i+1,"corriente":c,"media":media,"motivo":f"0.0 A (media:{media:.1f}A)"})
        elif c < umbral_v:
            res.append({"string":i+1,"corriente":c,"media":media,
                        "motivo":f"{c:.1f}A — {c/media*100:.0f}% de media({media:.1f}A)"})
    return res

def escanear_gateway(ip, id_start, id_end, opts=None):
    opts = opts or {}
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(TIMEOUT_MS/1000); s.connect((ip, PUERTO)); s.close()
    except:
        return [{"id":"-","estado":"SIN CONEXION","anomalias":[],"alarmas":{},
                 "diag":[],"voltaje":None}]
    resultados = []
    for unit in range(id_start, id_end+1):
        intento = 0; ok = False
        while intento < REINTENTOS and not ok:
            try:
                regs = modbus_read(ip, PUERTO, unit, REG_STATUS, 1, TIMEOUT_MS/1000)
                raw  = regs[0]
                bit5 = (raw>>5)&1
                estado = "COMUNICACION CORRECTA" if bit5 else "SIN COMUNICACION"
                alarmas = {}
                if opts.get("alarmas") and estado=="COMUNICACION CORRECTA":
                    if not ((raw>>0)&1): alarmas["seccionador"] = True
                    if not ((raw>>1)&1): alarmas["sobretension"] = True
                diag = []
                if opts.get("diag") and estado=="COMUNICACION CORRECTA":
                    diag = analizar_diagnostico(raw)
                voltaje = None
                if opts.get("voltaje") and estado=="COMUNICACION CORRECTA":
                    voltaje = leer_voltaje(ip, unit, TIMEOUT_MS/1000)
                anomalias = []
                if opts.get("strings") and estado=="COMUNICACION CORRECTA":
                    corr = leer_strings_caja(ip, unit, TIMEOUT_MS/1000)
                    anomalias = analizar_anomalias(corr, opts.get("umbral",30),
                                                   opts.get("strings_map",{}).get(unit))
                resultados.append({"id":unit,"estado":estado,"anomalias":anomalias,
                                   "alarmas":alarmas,"diag":diag,"voltaje":voltaje})
                ok = True
            except:
                intento += 1
                if intento >= REINTENTOS:
                    resultados.append({"id":unit,"estado":"ERROR","anomalias":[],
                                       "alarmas":{},"diag":[],"voltaje":None})
                else: time.sleep(DELAY_REINTENTO)
    return resultados

# ═══════════════════════════════════════════════════════════════
#  CBT — {gateway_ip: {modbus_id: mac}}
# ═══════════════════════════════════════════════════════════════
def parsear_cbt(ruta, gateway_ip):
    """Parsea un archivo CBT y lo asocia a una IP gateway.
    Devuelve {modbus_id: mac}"""
    macs = {}
    try:
        with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
            for linea in f:
                p = linea.strip().split()
                if len(p) < 2: continue
                try:
                    mid = int(p[0])
                    mac = p[1].strip().upper()
                    if len(mac)==16 and all(c in '0123456789ABCDEF' for c in mac):
                        macs[mid] = mac
                except: continue
    except: pass
    return macs

def mac_desde_cbt(cbt_por_ip, gateway_ip, modbus_id):
    """Busca MAC en el dict {ip: {mid: mac}}"""
    return (cbt_por_ip or {}).get(gateway_ip, {}).get(modbus_id)

# ═══════════════════════════════════════════════════════════════
#  TRAZABILIDAD EXCEL
# ═══════════════════════════════════════════════════════════════
def normalizar(texto):
    if texto is None: return ""
    t = str(texto).lower().strip()
    for o,r in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n"),
                ("à","a"),("â","a"),("ä","a"),("è","e"),("ê","e"),("ë","e"),
                ("ì","i"),("î","i"),("ï","i"),("ò","o"),("ô","o"),("ö","o"),
                ("ù","u"),("û","u"),("ü","u"),("ç","c")]:
        t = t.replace(o,r)
    return "".join(c for c in t if c.isalnum())

def detectar_col(cabs_norm, claves):
    kn = [normalizar(c) for c in claves]
    for k in kn:
        for i,c in enumerate(cabs_norm):
            if c and c == k: return i
    for k in kn:
        if len(k) < 2: continue
        for i,c in enumerate(cabs_norm):
            if c and (k in c or c in k): return i
    return None

def cruzar_trazabilidad(ruta_excel, resultados_por_gw):
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    except Exception as ex:
        return None, None, str(ex)
    mejor = col_ip = col_id = col_caja = col_str = None
    nombres = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        cab = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not cab: continue
        cn = [normalizar(c) for c in cab]
        ii = detectar_col(cn, CLAVES_IP)
        di = detectar_col(cn, CLAVES_ID)
        ci = detectar_col(cn, CLAVES_CAJA)
        si = detectar_col(cn, CLAVES_STRINGS)
        if ii is not None and di is not None:
            mejor=ws; col_ip=ii; col_id=di; col_caja=ci; col_str=si
            nombres={"ip":str(cab[ii]),"id":str(cab[di]),
                     "caja":str(cab[ci]) if ci is not None else "No detectada",
                     "strings":str(cab[si]) if si is not None else "No detectada"}
            break
    if not mejor: return None, None, "No se encontraron columnas IP e ID"
    def nip(v): return "".join(c for c in str(v).strip() if c.isdigit()) if v else ""
    def nid(v):
        try: return str(int(float(str(v).strip())))
        except: return normalizar(str(v))
    tabla = []
    for row in mejor.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row): continue
        ip_v  = nip(row[col_ip]) if col_ip is not None and col_ip < len(row) else ""
        id_v  = nid(row[col_id]) if col_id is not None and col_id < len(row) else ""
        caja  = row[col_caja] if col_caja is not None and col_caja < len(row) else None
        str_v = row[col_str]  if col_str  is not None and col_str  < len(row) else None
        ns = None
        if str_v is not None:
            try: ns = int(float(str(str_v).strip()))
            except: pass
        if ip_v or id_v:
            tabla.append({"ip":ip_v,"id":id_v,"caja":str(caja).strip() if caja else "","ns":ns})
    mapeo = []
    for ip, res in resultados_por_gw.items():
        ip_n = nip(ip)
        for r in res:
            if r["id"]=="-": continue
            id_n = nid(str(r["id"]))
            caja = ns = None
            for f in tabla:
                if f["ip"]==ip_n and f["id"]==id_n:
                    caja=f["caja"]; ns=f["ns"]; break
            mapeo.append({"ip":ip,"id":r["id"],"numero_caja":caja,"num_strings":ns})
    return mapeo, nombres, None

# ═══════════════════════════════════════════════════════════════
#  ESTADISTICAS
# ═══════════════════════════════════════════════════════════════
def calcular_stats(res_por_gw, orden=None):
    ips = orden or list(res_por_gw.keys())
    total=correcta=sin_com=error=sin_cx=0
    for ip in ips:
        for r in (res_por_gw.get(ip) or []):
            total+=1
            if r["id"]=="-": sin_cx+=1
            elif r["estado"]=="COMUNICACION CORRECTA": correcta+=1
            elif r["estado"]=="SIN COMUNICACION": sin_com+=1
            elif r["estado"]=="ERROR": error+=1
            elif r["estado"]=="SIN CONEXION": sin_cx+=1
    pok  = correcta/total*100 if total else 0
    pfail= (sin_com+error+sin_cx)/total*100 if total else 0
    return {"total":total,"correcta":correcta,"sin_com":sin_com,
            "error":error,"sin_conexion":sin_cx,"pct_correcta":pok,"pct_fallo":pfail}

# ═══════════════════════════════════════════════════════════════
#  PDF
# ═══════════════════════════════════════════════════════════════
AW, AH = A4
CA = colors.HexColor
CF = {"dark":"#0f172a","med":"#1e293b","border":"#334155","primary":"#0ea5e9",
      "green":"#16a34a","red":"#dc2626","yellow":"#d97706","orange":"#ea580c",
      "purple":"#7c3aed","gtext":"#475569","gclear":"#f1f5f9","gmed":"#e2e8f0"}

def _hf(cv, doc, titulo, fecha, portada=False):
    cv.saveState()
    if not portada:
        ch=1.2*cm; yc=AH-ch-0.7*cm
        cv.setFillColor(CA(CF["dark"]))
        cv.roundRect(doc.leftMargin-0.3*cm,yc-0.1*cm,
                     AW-doc.leftMargin-doc.rightMargin+0.6*cm,ch+0.2*cm,4,fill=1,stroke=0)
        cv.setFillColor(CA(CF["primary"]))
        cv.rect(doc.leftMargin-0.3*cm,yc+ch+0.1*cm,
                AW-doc.leftMargin-doc.rightMargin+0.6*cm,0.15*cm,fill=1,stroke=0)
        cv.setFillColor(colors.white); cv.setFont("Helvetica-Bold",10)
        cv.drawString(doc.leftMargin, yc+0.38*cm, "WEBDOM MODBUS SCANNER")
        if titulo:
            cv.setFillColor(CA("#94a3b8")); cv.setFont("Helvetica",8)
            cv.drawCentredString(AW/2, yc+0.40*cm, titulo)
        cv.setFillColor(CA("#64748b")); cv.setFont("Helvetica",8)
        cv.drawRightString(AW-doc.rightMargin, yc+0.40*cm, fecha)
    py=0.55*cm
    cv.setStrokeColor(CA(CF["border"])); cv.setLineWidth(0.4)
    cv.line(doc.leftMargin,py+0.5*cm,AW-doc.rightMargin,py+0.5*cm)
    cv.setFillColor(CA(CF["gtext"])); cv.setFont("Helvetica",7)
    cv.setFont("Helvetica-Bold",7.5); cv.setFillColor(CA(CF["primary"]))
    cv.drawRightString(AW-doc.rightMargin,py+0.15*cm,f"Pagina {doc.page}")
    cv.restoreState()

EST = [
    ("FONTSIZE",(0,0),(-1,-1),7),("ALIGN",(0,0),(-1,-1),"CENTER"),
    ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),5),
    ("BOTTOMPADDING",(0,0),(-1,-1),5),("LEFTPADDING",(0,0),(-1,-1),6),
    ("RIGHTPADDING",(0,0),(-1,-1),6),
    ("BACKGROUND",(0,0),(-1,0),CA(CF["med"])),
    ("TEXTCOLOR",(0,0),(-1,0),colors.white),
    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
    ("FONTSIZE",(0,0),(-1,0),7.5),
    ("ROWBACKGROUNDS",(0,1),(-1,-1),[CA(CF["gclear"]),CA(CF["gmed"])]),
    ("LINEBELOW",(0,0),(-1,0),1.2,CA(CF["primary"])),
    ("GRID",(0,1),(-1,-1),0.35,CA("#cbd5e1")),
    ("BOX",(0,0),(-1,-1),0.8,CA(CF["border"])),
]

def generar_pdf(res_gw, orden=None, mapeo=None, cols_det=None, ruta=None,
                titulo_proyecto="", opts=None, cbt_por_ip=None):
    opts = opts or {}
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    titulo = titulo_proyecto.strip() or "Sin nombre"

    # Nombre del archivo: titulo + fecha
    if ruta is None:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"{titulo}_{fecha_archivo}.pdf".replace(" ", "_").replace("/","")
        ruta = os.path.join(desktop, nombre_archivo)

    stats = calcular_stats(res_gw, orden)
    fallos = stats["sin_com"]+stats["error"]+stats["sin_conexion"]
    ips = orden or list(res_gw.keys())
    con_caja = bool(mapeo)
    lookup = {}
    if con_caja:
        for item in mapeo:
            lookup[(str(item["ip"]),str(item["id"]))] = item.get("numero_caja") or "—"

    doc = BaseDocTemplate(ruta, pagesize=A4,
                          leftMargin=2*cm, rightMargin=2*cm,
                          topMargin=2.8*cm, bottomMargin=1.8*cm)
    aw = AW-doc.leftMargin-doc.rightMargin
    fr_port = Frame(doc.leftMargin,doc.bottomMargin,aw,AH-doc.bottomMargin-1.5*cm,id="p")
    fr_norm = Frame(doc.leftMargin,doc.bottomMargin,aw,AH-doc.topMargin-doc.bottomMargin,id="n")
    doc.addPageTemplates([
        PageTemplate(id="portada",frames=[fr_port],onPage=lambda c,d:_hf(c,d,titulo,fecha,True)),
        PageTemplate(id="normal", frames=[fr_norm],onPage=lambda c,d:_hf(c,d,titulo,fecha,False)),
    ])
    st = getSampleStyleSheet()
    def _s(nm,**kw): return ParagraphStyle(nm,parent=st["Normal"],**kw)
    sh1=_s("h1",fontSize=30,fontName="Helvetica-Bold",textColor=CA(CF["dark"]),alignment=1,spaceAfter=6)
    sh2=_s("h2",fontSize=11,fontName="Helvetica-Bold",textColor=CA(CF["primary"]),spaceBefore=14,spaceAfter=5)
    sh3=_s("h3",fontSize=9.5,fontName="Helvetica-Bold",textColor=CA(CF["border"]),spaceBefore=8,spaceAfter=3)
    sn =_s("n", fontSize=8.5,textColor=CA(CF["gtext"]))
    sp =_s("p", fontSize=10,fontName="Helvetica",textColor=CA(CF["gtext"]),alignment=1)
    sn2=_s("n2",fontSize=8,fontName="Helvetica-Oblique",textColor=CA("#94a3b8"),alignment=1)

    elems = []
    # ── PORTADA ──
    kw = aw/3
    def kpi(lbl,val,sub,col):
        return Paragraph(
            f'<font size="7" color="#64748b"><b>{lbl}</b></font><br/><br/>'
            f'<font size="28" color="{col}"><b>{val}</b></font><br/>'
            f'<font size="10" color="#64748b">{sub}</font>',
            _s(f"kp{lbl[:2]}",alignment=1,leading=16))

    elems += [Spacer(1,3.5*cm),
              HRFlowable(width="100%",thickness=2.5,color=CA(CF["primary"]),spaceAfter=14),
              Paragraph("WEBDOM MODBUS SCANNER",sh1),
              Spacer(1,0.4*cm),
              HRFlowable(width="60%",thickness=0.8,color=CA(CF["border"]),spaceAfter=10),
              Spacer(1,0.2*cm),
              Paragraph(f"Proyecto: {titulo}",_s("sub",fontSize=18,fontName="Helvetica-Bold",
                        textColor=CA(CF["dark"]),alignment=1,spaceAfter=14)),
              Paragraph(f"Generado el {datetime.now().strftime('%d de %B de %Y')} a las {datetime.now().strftime('%H:%M')} h",sp),
              Spacer(1,1.4*cm),
              HRFlowable(width="100%",thickness=0.8,color=CA(CF["border"]),spaceAfter=12)]

    kt=Table([[kpi("TOTAL CAJAS",str(stats["total"]),"escaneadas","#0f172a"),
               kpi("COMUNICAN OK",str(stats["correcta"]),f"{stats['pct_correcta']:.1f}%","#16a34a"),
               kpi("CON FALLO",str(fallos),f"{stats['pct_fallo']:.1f}%","#dc2626")]],
             colWidths=[kw]*3,rowHeights=[3.8*cm])
    kt.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),14),("BOTTOMPADDING",(0,0),(-1,-1),14),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("LINEAFTER",(0,0),(1,-1),0.6,CA(CF["border"])),
        ("BOX",(0,0),(-1,-1),0.8,CA(CF["border"])),
        ("BACKGROUND",(0,0),(-1,-1),CA(CF["gclear"])),
    ]))
    elems.append(kt)
    elems.append(Spacer(1,0.5*cm))
    elems.append(Paragraph(
        f"Sin conexion: {stats['sin_conexion']}  |  Error lectura: {stats['error']}  |  Sin com Modbus: {stats['sin_com']}",
        _s("d",fontSize=8.5,textColor=CA(CF["gtext"]),alignment=1)))

    if con_caja and cols_det:
        elems.append(Spacer(1,0.4*cm))
        elems.append(Paragraph(
            f"Trazabilidad — IP:'{cols_det.get('ip','?')}' | ID:'{cols_det.get('id','?')}' | "
            f"Caja:'{cols_det.get('caja','?')}' | Strings:'{cols_det.get('strings','No detectada')}'",sn2))

    elems += [Spacer(1,0.7*cm),
              HRFlowable(width="100%",thickness=2.5,color=CA(CF["primary"]),spaceAfter=6),
              PageBreak()]

    # ── SECCION 1 — DETALLE ──
    CE={"COMUNICACION CORRECTA":CA(CF["green"]),"SIN COMUNICACION":CA(CF["red"]),
        "ERROR":CA(CF["red"]),"SIN CONEXION":CA(CF["purple"])}
    elems.append(Paragraph("1. DETALLE POR GATEWAY",sh2))
    elems.append(HRFlowable(width="100%",thickness=0.8,color=CA(CF["primary"]),spaceAfter=8))

    usar_mac      = opts.get("mac",False)
    usar_volt     = opts.get("voltaje",False)
    usar_alarm    = opts.get("alarmas",False)
    usar_diag     = opts.get("diag",False)
    usar_strings  = opts.get("strings",False)

    for ip in ips:
        res = res_gw.get(ip)
        if not res: continue
        elems.append(Paragraph(f"Gateway: {ip}",sh3))

        # Cabecera dinamica
        if con_caja:
            cab=["N.CAJA","GW IP","ID","ESTADO"]
            cw=[2.8*cm,3.5*cm,1.8*cm,4.5*cm]
        else:
            cab=["GW IP","ID","ESTADO"]
            cw=[4.0*cm,2.0*cm,6.0*cm]
        if usar_alarm: cab.append("ALARMAS"); cw.append(2.5*cm)
        if usar_diag:  cab.append("DIAG");    cw.append(2.2*cm)
        if usar_volt:  cab.append("V DC");    cw.append(1.6*cm)
        if usar_mac:   cab.append("MAC");     cw.append(4.0*cm)
        # Ajustar ultima columna para llenar ancho
        total_cw = sum(cw)
        if total_cw < aw - 0.5*cm:
            cw[-1] += (aw - 0.5*cm - total_cw)

        data=[cab]; cmds=list(EST)
        idx_est = cab.index("ESTADO")
        idx_alr = cab.index("ALARMAS") if usar_alarm else None
        for i,r in enumerate(res,1):
            if not isinstance(r.get("id"),int): continue
            caja = lookup.get((str(ip),str(r["id"])),"—") if con_caja else None
            est = r["estado"]
            row = ([str(caja),str(ip),str(r["id"]),est] if con_caja
                   else [str(ip),str(r["id"]),est])
            c_est = CE.get(est)
            if c_est:
                cmds.append(("TEXTCOLOR",(idx_est,i),(idx_est,i),c_est))
                cmds.append(("FONTNAME",(idx_est,i),(idx_est,i),"Helvetica-Bold"))
            if usar_alarm:
                alm=r.get("alarmas",{})
                al_txt=[]
                if alm.get("seccionador"): al_txt.append("SECC.")
                if alm.get("sobretension"): al_txt.append("SOBRETEN.")
                cel = " / ".join(al_txt) if al_txt else ("OK" if est=="COMUNICACION CORRECTA" else "—")
                row.append(cel)
                if al_txt:
                    cmds.append(("TEXTCOLOR",(idx_alr,i),(idx_alr,i),CA(CF["orange"])))
                    cmds.append(("FONTNAME",(idx_alr,i),(idx_alr,i),"Helvetica-Bold"))
                elif est=="COMUNICACION CORRECTA":
                    cmds.append(("TEXTCOLOR",(idx_alr,i),(idx_alr,i),CA(CF["green"])))
            if usar_diag:
                d=r.get("diag",[])
                row.append(" / ".join(d) if d else ("" if est=="COMUNICACION CORRECTA" else "—"))
            if usar_volt:
                v=r.get("voltaje")
                row.append(f"{v:.0f}V" if v is not None else "—")
            if usar_mac:
                # Buscar MAC: primero en CBT por IP, luego fallback
                mac = mac_desde_cbt(cbt_por_ip, ip, r["id"])
                row.append(mac or "—")
            data.append(row)
        if len(data)>1:
            t=Table(data,colWidths=cw,repeatRows=1)
            t.setStyle(TableStyle(cmds))
            elems.append(t); elems.append(Spacer(1,0.4*cm))

    # ── SECCION 2 — INCIDENCIAS ──
    elems += [Spacer(1,0.3*cm),
              Paragraph("2. RESUMEN DE INCIDENCIAS",sh2),
              HRFlowable(width="100%",thickness=0.8,color=CA(CF["red"]),spaceAfter=8),
              Paragraph(f"Total: {stats['total']} cajas | Correctas: {stats['correcta']} ({stats['pct_correcta']:.1f}%) | Fallos: {fallos} ({stats['pct_fallo']:.1f}%)",sn),
              Spacer(1,0.3*cm)]

    if con_caja:
        cab2=["GW IP","ID","N.CAJA","ESTADO"]; cw2=[4.0*cm,2.0*cm,3.0*cm,6.2*cm]
    else:
        cab2=["GW IP","ID","ESTADO"]; cw2=[4.5*cm,2.5*cm,8.2*cm]
    data2=[cab2]; hay=False
    for ip in ips:
        for r in (res_gw.get(ip) or []):
            if r["estado"]!="COMUNICACION CORRECTA":
                hay=True
                caja=lookup.get((str(ip),str(r["id"])),"—") if con_caja else None
                row2=([str(ip),str(r["id"]),str(caja),r["estado"]] if con_caja
                      else [str(ip),str(r["id"]),r["estado"]])
                data2.append(row2)
    if hay:
        ie=[("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),("LEFTPADDING",(0,0),(-1,-1),6),
            ("RIGHTPADDING",(0,0),(-1,-1),6),
            ("BACKGROUND",(0,0),(-1,0),CA("#7f1d1d")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8.5),
            ("LINEBELOW",(0,0),(-1,0),1.2,CA("#f87171")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[CA("#fff1f2"),CA("#ffe4e6")]),
            ("GRID",(0,1),(-1,-1),0.35,CA("#fca5a5")),("BOX",(0,0),(-1,-1),0.8,CA("#ef4444"))]
        ti=Table(data2,colWidths=cw2,repeatRows=1)
        ti.setStyle(TableStyle(ie)); elems.append(ti)
    else:
        elems.append(Paragraph("Sin incidencias. Todos los dispositivos comunican correctamente.",sn))

    # ── SECCION 3 — ALARMAS ──
    if usar_alarm:
        cajas_alm=[]
        for ip in ips:
            for r in (res_gw.get(ip) or []):
                alm=r.get("alarmas",{})
                if alm:
                    caja=lookup.get((str(ip),str(r["id"])),"—") if con_caja else "—"
                    cajas_alm.append({"ip":ip,"id":r["id"],"caja":caja,"alarmas":alm})
        ns=sum(1 for x in cajas_alm if x["alarmas"].get("seccionador"))
        nb=sum(1 for x in cajas_alm if x["alarmas"].get("sobretension"))
        elems += [Spacer(1,0.4*cm),
                  Paragraph("3. ALARMAS: SECCIONADOR Y SOBRETENSION",sh2),
                  HRFlowable(width="100%",thickness=0.8,color=CA(CF["orange"]),spaceAfter=8),
                  Paragraph(f"Seccionador abierto: {ns} | Sobretension: {nb} | Total con alarma: {len(cajas_alm)}",sn),
                  Spacer(1,0.3*cm)]
        if cajas_alm:
            if con_caja:
                ca2=["GW IP","ID","N.CAJA","SECCIONADOR","SOBRETENSION"]
                cwa=[3.5*cm,1.8*cm,3.0*cm,3.2*cm,3.5*cm]
            else:
                ca2=["GW IP","ID","SECCIONADOR","SOBRETENSION"]
                cwa=[4.5*cm,2.0*cm,3.5*cm,3.5*cm]
            da=[ca2]
            ae=[("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),5),
                ("BOTTOMPADDING",(0,0),(-1,-1),5),("LEFTPADDING",(0,0),(-1,-1),6),
                ("RIGHTPADDING",(0,0),(-1,-1),6),
                ("BACKGROUND",(0,0),(-1,0),CA("#7c2d12")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8.5),
                ("LINEBELOW",(0,0),(-1,0),1.2,CA("#fb923c")),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[CA("#fff7ed"),CA("#ffedd5")]),
                ("GRID",(0,1),(-1,-1),0.35,CA("#fdba74")),
                ("BOX",(0,0),(-1,-1),0.8,CA("#ea580c")),
                ("TEXTCOLOR",(0,1),(-1,-1),CA("#9a3412")),
                ("FONTNAME",(0,1),(-1,-1),"Helvetica-Bold")]
            is_=ca2.index("SECCIONADOR"); ib_=ca2.index("SOBRETENSION")
            for i,item in enumerate(cajas_alm,1):
                s_=("⚠ ABIERTO" if item["alarmas"].get("seccionador") else "OK")
                b_=("⚠ ALARMA"  if item["alarmas"].get("sobretension") else "OK")
                row3=([str(item["ip"]),str(item["id"]),str(item["caja"]),s_,b_] if con_caja
                      else [str(item["ip"]),str(item["id"]),s_,b_])
                da.append(row3)
                ae.append(("TEXTCOLOR",(is_,i),(is_,i),CA(CF["orange"]) if item["alarmas"].get("seccionador") else CA(CF["green"])))
                ae.append(("TEXTCOLOR",(ib_,i),(ib_,i),CA(CF["orange"]) if item["alarmas"].get("sobretension") else CA(CF["green"])))
            ta=Table(da,colWidths=cwa,repeatRows=1); ta.setStyle(TableStyle(ae))
            elems.append(ta)
        else:
            elems.append(Paragraph("No se han detectado alarmas.",sn))

    # ── SECCION — STRINGS ──
    sn_num = 4 if usar_alarm else 3
    if usar_strings:
        hay_a=any(any(r.get("anomalias") for r in (res_gw.get(ip) or [])) for ip in ips)
        if hay_a:
            elems += [Spacer(1,0.4*cm),
                      Paragraph(f"{sn_num}. ANOMALIAS DE STRINGS",sh2),
                      HRFlowable(width="100%",thickness=0.8,color=CA(CF["yellow"]),spaceAfter=8)]
            for ip in ips:
                for r in (res_gw.get(ip) or []):
                    an=r.get("anomalias") or []
                    if not an: continue
                    caja_lbl=""
                    if con_caja:
                        caja_lbl=f"  —  Caja: {lookup.get((str(ip),str(r['id'])),'—')}"
                    elems.append(Paragraph(f"GW {ip} | ID {r['id']}{caja_lbl}",sh3))
                    dan=[["STRING","CORRIENTE","DESCRIPCION"]]
                    for a in an:
                        dan.append([f"S{a['string']}",f"{a['corriente']:.1f}A",a["motivo"]])
                    ta=Table(dan,colWidths=[2*cm,3*cm,aw-5.5*cm])
                    ae2=[("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(1,-1),"CENTER"),
                         ("ALIGN",(2,0),(2,-1),"LEFT"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                         ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
                         ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
                         ("BACKGROUND",(0,0),(-1,0),CA("#78350f")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
                         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8.5),
                         ("LINEBELOW",(0,0),(-1,0),1.2,CA("#fcd34d")),
                         ("ROWBACKGROUNDS",(0,1),(-1,-1),[CA("#fffbeb"),CA("#fef3c7")]),
                         ("GRID",(0,1),(-1,-1),0.35,CA("#fde68a")),("BOX",(0,0),(-1,-1),0.8,CA("#f59e0b")),
                         ("TEXTCOLOR",(0,1),(-1,-1),CA("#92400e")),("FONTNAME",(0,1),(-1,-1),"Helvetica-Bold")]
                    ta.setStyle(TableStyle(ae2)); elems.append(ta); elems.append(Spacer(1,0.3*cm))

    doc.build(elems)
    return ruta

# ═══════════════════════════════════════════════════════════════
#  CSV
# ═══════════════════════════════════════════════════════════════
def generar_csv(res_gw, orden=None, mapeo=None, ruta=None, titulo_proyecto="",
                opts=None, cbt_por_ip=None):
    opts = opts or {}
    titulo = titulo_proyecto.strip() or "Escaneo"
    if ruta is None:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        fecha_a = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_a = f"{titulo}_{fecha_a}.csv".replace(" ","_").replace("/","")
        ruta = os.path.join(desktop, nombre_a)

    con_caja = bool(mapeo)
    lookup = {}
    if con_caja:
        for item in mapeo:
            lookup[(str(item["ip"]),str(item["id"]))] = item.get("numero_caja") or ""
    ips = orden or list(res_gw.keys())
    usar_mac   = opts.get("mac",False)
    usar_volt  = opts.get("voltaje",False)
    usar_alarm = opts.get("alarmas",False)
    usar_diag  = opts.get("diag",False)
    usar_str   = opts.get("strings",False)

    buf=io.StringIO(); wr=csv.writer(buf,delimiter=";")
    cab=[]
    if con_caja: cab.append("N_Caja")
    cab+=["Gateway_IP","ID_Modbus","Estado"]
    if usar_alarm: cab+=["Alarma_Secc","Alarma_Sobreten"]
    if usar_diag:  cab.append("Diagnostico")
    if usar_volt:  cab.append("Voltaje_V")
    if usar_mac:   cab.append("MAC")
    if usar_str:   cab+=[f"S{i}" for i in range(1,NUM_STRINGS+1)]
    wr.writerow(cab)

    for ip in ips:
        for r in (res_gw.get(ip) or []):
            if not isinstance(r.get("id"),int): continue
            row=[]
            if con_caja: row.append(lookup.get((str(ip),str(r["id"])),""))
            row+=[str(ip),str(r["id"]),r["estado"]]
            if usar_alarm:
                alm=r.get("alarmas",{})
                row.append("ALARMA" if alm.get("seccionador") else ("OK" if r["estado"]=="COMUNICACION CORRECTA" else "—"))
                row.append("ALARMA" if alm.get("sobretension") else ("OK" if r["estado"]=="COMUNICACION CORRECTA" else "—"))
            if usar_diag:
                d=r.get("diag",[])
                row.append(" / ".join(d) if d else "")
            if usar_volt:
                v=r.get("voltaje")
                row.append(f"{v:.0f}" if v is not None else "")
            if usar_mac:
                mac=mac_desde_cbt(cbt_por_ip,ip,r["id"])
                row.append(mac or "")
            if usar_str:
                am={a["string"]:a["motivo"] for a in (r.get("anomalias") or [])}
                for s in range(1,NUM_STRINGS+1): row.append(am.get(s,""))
            wr.writerow(row)

    with open(ruta,"w",encoding="utf-8-sig",newline="") as f: f.write(buf.getvalue())
    return ruta

# ═══════════════════════════════════════════════════════════════
#  APP FLET
# ═══════════════════════════════════════════════════════════════
def main(page: ft.Page):
    init_db()
    page.title = "Webdom Modbus Scanner"
    page.theme_mode = ft.ThemeMode.DARK
    page.bgcolor = "#0f172a"
    page.scroll = ft.ScrollMode.AUTO
    page.window.width = 860
    page.window.min_width = 720

    gateways     = []
    res_globales = {}
    excel_ref    = {"path": None}
    # CBT por IP: {"10.0.1.1": {1: "AABB...", 2: "CCDD..."}, ...}
    cbt_por_ip   = {}
    cancel_ref   = {"v": False}
    planta_activa= {"id": None}  # ID de la planta cargada

    # ── Widgets ──
    txt_ip    = ft.TextField(label="IP Gateway", width=240, border_color="#334155", focused_border_color="#38bdf8")
    txt_ini   = ft.TextField(label="ID inicio",  width=110, border_color="#334155", focused_border_color="#38bdf8")
    txt_fin   = ft.TextField(label="ID fin",     width=110, border_color="#334155", focused_border_color="#38bdf8")
    txt_proy  = ft.TextField(label="Nombre del proyecto / reporte", width=420,
                             border_color="#334155", focused_border_color="#a78bfa",
                             hint_text="Ej: Planta Solar Norte")

    lista_gw  = ft.Column(spacing=4)
    prog      = ft.ProgressBar(value=0, color="#38bdf8", bgcolor="#1e293b")
    lbl_est   = ft.Text("Listo", color="#94a3b8", size=13)
    lbl_excel = ft.Text("Sin Excel", color="#64748b", size=12, italic=True)
    lbl_cbt   = ft.Text("Sin CBT cargado", color="#64748b", size=12, italic=True)
    lbl_planta= ft.Text("Sin planta cargada", color="#64748b", size=11, italic=True)

    lbl_tot   = ft.Text("—", size=22, weight="bold", color="#f1f5f9")
    lbl_ok    = ft.Text("—", size=22, weight="bold", color="#4ade80")
    lbl_ko    = ft.Text("—", size=22, weight="bold", color="#f87171")
    lbl_pok   = ft.Text("—", size=14, color="#4ade80")
    lbl_pko   = ft.Text("—", size=14, color="#f87171")

    chk_strings = ft.Checkbox(label="Anomalias strings", value=False, active_color="#f59e0b")
    txt_umbral  = ft.TextField(label="Umbral %", value="30", width=110,
                               border_color="#334155", focused_border_color="#f59e0b")
    chk_alarmas = ft.Checkbox(label="Alarmas (secc/sobreten)", value=False, active_color="#ea580c")
    chk_diag    = ft.Checkbox(label="Diagnostico (LoRa/Mem)", value=False, active_color="#ef4444")
    chk_voltaje = ft.Checkbox(label="Voltaje DC", value=False, active_color="#facc15")
    chk_mac     = ft.Checkbox(label="MAC (desde CBT)", value=False, active_color="#8b5cf6")

    # Lista de plantas guardadas
    plantas_dropdown = ft.Dropdown(label="Plantas guardadas", width=300,
                                   border_color="#334155", focused_border_color="#38bdf8",
                                   options=[], hint_text="Selecciona una planta")

    tabla = ft.DataTable(
        border=ft.border.all(1,"#1e293b"), border_radius=8,
        heading_row_color="#1e293b", data_row_min_height=36,
        columns=[
            ft.DataColumn(ft.Text("Gateway",    color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("ID",         color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("Estado",     color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("Strings",    color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("Alarmas",    color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("Diagnostico",color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("Voltaje DC", color="#94a3b8", weight="bold")),
            ft.DataColumn(ft.Text("MAC",        color="#94a3b8", weight="bold")),
        ], rows=[])

    seccion_stats = ft.Container(
        content=ft.Column([
            ft.Text("RESUMEN", size=11, weight="bold", color="#475569"),
            ft.Row([
                ft.Container(content=ft.Column([ft.Text("TOTAL",size=9,color="#475569",weight="bold"),lbl_tot],
                    spacing=2,horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                    bgcolor="#0f172a",border=ft.border.all(1,"#334155"),border_radius=8,
                    padding=ft.padding.symmetric(10,14),expand=True),
                ft.Container(content=ft.Column([ft.Text("COMUNICAN",size=9,color="#4ade80",weight="bold"),lbl_ok,lbl_pok],
                    spacing=2,horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                    bgcolor="#0f172a",border=ft.border.all(1,"#16a34a"),border_radius=8,
                    padding=ft.padding.symmetric(10,14),expand=True),
                ft.Container(content=ft.Column([ft.Text("FALLOS",size=9,color="#f87171",weight="bold"),lbl_ko,lbl_pko],
                    spacing=2,horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                    bgcolor="#0f172a",border=ft.border.all(1,"#dc2626"),border_radius=8,
                    padding=ft.padding.symmetric(10,14),expand=True),
            ],spacing=8),
        ],spacing=8),
        bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"),
        border_radius=10, padding=14, visible=False)

    def actualizar_stats(stats):
        lbl_tot.value = str(stats["total"])
        lbl_ok.value  = str(stats["correcta"])
        lbl_ko.value  = str(stats["sin_com"]+stats["error"]+stats["sin_conexion"])
        lbl_pok.value = f"{stats['pct_correcta']:.1f}%"
        lbl_pko.value = f"{stats['pct_fallo']:.1f}%"
        page.update()

    # ── Gestión de gateways en UI ──
    def refresh_gws():
        lista_gw.controls.clear()
        for idx,(ip,s,en) in enumerate(gateways):
            def mk_rm(i):
                def fn(e): gateways.pop(i); refresh_gws(); page.update()
                return fn
            lista_gw.controls.append(ft.Container(
                content=ft.Row([
                    ft.Icon(ft.icons.ROUTER, color="#38bdf8", size=16),
                    ft.Text(f"{ip}  IDs {s}–{en}", color="#e2e8f0", size=13, expand=True),
                    ft.IconButton(ft.icons.DELETE_OUTLINE, icon_color="#f87171", icon_size=16,
                                  on_click=mk_rm(idx),
                                  style=ft.ButtonStyle(padding=ft.padding.all(4))),
                ],spacing=6),
                bgcolor="#1e293b", border_radius=6, padding=ft.padding.symmetric(6,10)))

    def add_gw(e):
        ip=txt_ip.value.strip(); sv=txt_ini.value.strip(); ev=txt_fin.value.strip()
        if not ip or not sv or not ev: return
        try: s,en=int(sv),int(ev)
        except: return
        if en < s: return
        gateways.append((ip,s,en)); refresh_gws()
        txt_ip.value=txt_ini.value=txt_fin.value=""
        page.update()

    # ── Plantas guardadas ──
    def refresh_dropdown_plantas():
        ps = db_get_plantas()
        plantas_dropdown.options = [
            ft.dropdown.Option(str(p["id"]), p["nombre"]) for p in ps
        ]
        page.update()

    def on_cargar_planta(e):
        sel = plantas_dropdown.value
        if not sel: return
        pid = int(sel)
        ps = db_get_plantas()
        p = next((x for x in ps if x["id"]==pid), None)
        if not p: return
        gateways.clear()
        for gw in p["gateways"]:
            gateways.append((gw["ip"], gw["id_inicio"], gw["id_fin"]))
        refresh_gws()
        txt_proy.value = p["nombre"]
        planta_activa["id"] = pid
        lbl_planta.value = f"Planta: {p['nombre']} ({len(gateways)} GW)"
        lbl_planta.color = "#4ade80"
        # Cargar CBT guardada de esta planta
        cbt_db = db_get_cbt(pid)
        cbt_por_ip.clear()
        for (gw_ip, mid), mac in cbt_db.items():
            if gw_ip not in cbt_por_ip:
                cbt_por_ip[gw_ip] = {}
            cbt_por_ip[gw_ip][mid] = mac
        cnt_mac = sum(len(v) for v in cbt_por_ip.values())
        if cnt_mac:
            lbl_cbt.value = f"CBT: {cnt_mac} MACs (guardadas en planta)"
            lbl_cbt.color = "#4ade80"
        page.update()

    def on_guardar_planta(e):
        nombre = txt_proy.value.strip()
        if not nombre or not gateways:
            page.snack_bar = ft.SnackBar(ft.Text("Añade un nombre y al menos un gateway",color="white"),bgcolor="#dc2626")
            page.snack_bar.open=True; page.update(); return
        pid = db_guardar_planta(nombre, gateways)
        planta_activa["id"] = pid
        # Guardar también la CBT actual asociada a esta planta
        for gw_ip, macs in cbt_por_ip.items():
            db_guardar_cbt(pid, gw_ip, macs)
        lbl_planta.value = f"Planta '{nombre}' guardada"
        lbl_planta.color = "#4ade80"
        refresh_dropdown_plantas()
        page.snack_bar = ft.SnackBar(ft.Text(f"Planta '{nombre}' guardada",color="white"),bgcolor="#16a34a")
        page.snack_bar.open=True; page.update()

    def on_eliminar_planta(e):
        sel = plantas_dropdown.value
        if not sel: return
        pid = int(sel)
        db_eliminar_planta(pid)
        refresh_dropdown_plantas()
        page.snack_bar = ft.SnackBar(ft.Text("Planta eliminada",color="white"),bgcolor="#dc2626")
        page.snack_bar.open=True; page.update()

    refresh_dropdown_plantas()

    # ── CBT por IP ──
    cbt_ip_pendiente = {"ip": None}  # IP seleccionada para asignar el CBT

    def on_cbt_result(e: ft.FilePickerResultEvent):
        if not e.files: return
        ruta = e.files[0].path
        ip_asignar = cbt_ip_pendiente.get("ip") or ""
        if not ip_asignar:
            # Si no se especificó IP, usar la primera de los gateways
            ip_asignar = gateways[0][0] if gateways else "?"
        macs = parsear_cbt(ruta, ip_asignar)
        if ip_asignar not in cbt_por_ip:
            cbt_por_ip[ip_asignar] = {}
        cbt_por_ip[ip_asignar].update(macs)
        # Si hay planta activa, guardar en DB
        if planta_activa["id"]:
            db_guardar_cbt(planta_activa["id"], ip_asignar, macs)
        cnt = sum(len(v) for v in cbt_por_ip.values())
        lbl_cbt.value = f"CBT: {cnt} MACs ({len(cbt_por_ip)} IPs)"
        lbl_cbt.color = "#4ade80"
        page.update()

    fp_cbt = ft.FilePicker(on_result=on_cbt_result)

    # Dropdown para seleccionar qué IP se le asigna el CBT
    cbt_ip_dropdown = ft.Dropdown(label="IP del Gateway para este CBT", width=220,
                                  border_color="#334155", focused_border_color="#8b5cf6",
                                  hint_text="Selecciona IP", options=[])

    def on_cbt_ip_change(e):
        cbt_ip_pendiente["ip"] = cbt_ip_dropdown.value

    cbt_ip_dropdown.on_change = on_cbt_ip_change

    def refresh_cbt_ips():
        cbt_ip_dropdown.options = [ft.dropdown.Option(ip, ip) for ip,_,_ in gateways]
        if gateways: cbt_ip_dropdown.value = gateways[0][0]; cbt_ip_pendiente["ip"] = gateways[0][0]
        page.update()

    # ── Excel trazabilidad ──
    def on_excel(e: ft.FilePickerResultEvent):
        if e.files:
            excel_ref["path"] = e.files[0].path
            lbl_excel.value = os.path.basename(e.files[0].path)
            lbl_excel.color = "#4ade80"
        else:
            excel_ref["path"] = None
            lbl_excel.value = "Sin Excel"
            lbl_excel.color = "#64748b"
        page.update()

    fp_excel = ft.FilePicker(on_result=on_excel)
    page.overlay.append(fp_cbt)
    page.overlay.append(fp_excel)

    # ── ESCANEO ──
    def escanear(e):
        cancel_ref["v"] = False
        prog.value = 0
        lbl_est.value = "Escaneando..."
        tabla.rows.clear()
        res_globales.clear()
        seccion_stats.visible = False
        refresh_cbt_ips()
        page.update()

        opts = {
            "strings": chk_strings.value,
            "umbral":  int(txt_umbral.value or 30),
            "alarmas": chk_alarmas.value,
            "diag":    chk_diag.value,
            "voltaje": chk_voltaje.value,
            "mac":     chk_mac.value,
        }

        def worker():
            if not gateways:
                lbl_est.value = "Sin gateways"; page.update(); return

            # Leer num_strings del Excel si hay
            strings_map_global = {}
            if opts["strings"] and excel_ref["path"]:
                try:
                    lbl_est.value = "Leyendo strings del Excel..."; page.update()
                    mp,_,_ = cruzar_trazabilidad(excel_ref["path"],
                                                  {ip:[] for ip,_,_ in gateways})
                    if mp:
                        for item in mp:
                            if item.get("num_strings"):
                                if item["ip"] not in strings_map_global:
                                    strings_map_global[item["ip"]] = {}
                                try: strings_map_global[item["ip"]][int(item["id"])] = int(item["num_strings"])
                                except: pass
                except: pass

            n = len(gateways)
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
                def run_gw(gw):
                    ip,ini,fin = gw
                    o = dict(opts)
                    o["strings_map"] = strings_map_global.get(ip,{})
                    return ip, escanear_gateway(ip,ini,fin,o)
                futures = {ex.submit(run_gw,gw): gw for gw in gateways}
                done = 0
                for f in as_completed(futures):
                    if cancel_ref["v"]:
                        lbl_est.value="Cancelado"; page.update(); return
                    ip,res = f.result()
                    res_globales[ip] = res; done+=1
                    prog.value=done/n
                    lbl_est.value=f"Gateway {done}/{n}"; page.update()

            # Rellenar tabla
            tabla.rows.clear()
            for ip,ini,fin in gateways:
                for r in (res_globales.get(ip) or []):
                    ce = {"COMUNICACION CORRECTA":"#4ade80","SIN COMUNICACION":"#f87171",
                          "ERROR":"#f87171","SIN CONEXION":"#c084fc"}.get(r["estado"],"#94a3b8")

                    an=r.get("anomalias") or []
                    
                    an_lbl = f"⚠ {','.join('S' + str(a['string']) for a in an)}" if an else ("OK" if r["estado"]=="COMUNICACION CORRECTA" else "—") if opts["strings"] else "—"
                    an_col="#f59e0b" if an else ("#4ade80" if r["estado"]=="COMUNICACION CORRECTA" and opts["strings"] else "#475569")

                    alm=r.get("alarmas",{})
                    alr_p=[]; _has_alr=False
                    if alm.get("seccionador"): alr_p.append("⚠Secc"); _has_alr=True
                    if alm.get("sobretension"): alr_p.append("⚠Sobreten"); _has_alr=True
                    alr_lbl=" / ".join(alr_p) if alr_p else ("OK" if r["estado"]=="COMUNICACION CORRECTA" and opts["alarmas"] else "—")
                    alr_col="#fb923c" if _has_alr else ("#4ade80" if opts["alarmas"] and r["estado"]=="COMUNICACION CORRECTA" else "#475569")

                    dg=r.get("diag",[])
                    dg_lbl=" / ".join(f"⚠{d}" for d in dg) if dg else ("" if r["estado"]=="COMUNICACION CORRECTA" and opts["diag"] else ("—" if not opts["diag"] else "—"))
                    dg_col="#ef4444" if dg else "#475569"

                    v=r.get("voltaje")
                    v_lbl=f"{v:.0f}V" if v is not None else ("—" if not opts["voltaje"] else "N/D")
                    v_col=("#4ade80" if v and v>100 else "#f59e0b" if v and v>10 else "#f87171") if v is not None else "#475569"

                    # MAC: buscar en cbt_por_ip por IP del gateway y modbus_id
                    mac = mac_desde_cbt(cbt_por_ip, ip, r["id"]) if opts["mac"] else None
                    m_lbl = mac if mac else ("N/D" if opts["mac"] else "—")
                    m_col = "#a78bfa" if mac else "#475569"

                    row_col = ft.colors.with_opacity(0.08,"#fb923c") if _has_alr and r["estado"]=="COMUNICACION CORRECTA" else None

                    tabla.rows.append(ft.DataRow(color=row_col, cells=[
                        ft.DataCell(ft.Text(ip,              color="#e2e8f0")),
                        ft.DataCell(ft.Text(str(r["id"]),    color="#e2e8f0")),
                        ft.DataCell(ft.Text(r["estado"],     color=ce,     weight="bold")),
                        ft.DataCell(ft.Text(an_lbl,          color=an_col, weight="bold")),
                        ft.DataCell(ft.Text(alr_lbl,         color=alr_col,weight="bold")),
                        ft.DataCell(ft.Text(dg_lbl,          color=dg_col, weight="bold", size=11)),
                        ft.DataCell(ft.Text(v_lbl,           color=v_col,  weight="bold")),
                        ft.DataCell(ft.Text(m_lbl,           color=m_col,  size=11)),
                    ]))

            stats = calcular_stats(res_globales, [ip for ip,_,_ in gateways])
            actualizar_stats(stats)
            seccion_stats.visible = True
            lbl_est.value = "Completado"
            page.update()
            mostrar_exportar()

        threading.Thread(target=worker, daemon=True).start()

    def cancelar(e):
        cancel_ref["v"] = True
        lbl_est.value = "Cancelando..."; page.update()

    # ── EXPORTAR ──
    def mostrar_exportar():
        def do_exp(e2):
            opcion = grp.value
            usar_traz = chk_traz.value and excel_ref["path"]
            dlg.open = False; page.update()

            def run():
                mapeo = cols_det = None
                if usar_traz:
                    lbl_est.value = "Cruzando trazabilidad..."; page.update()
                    mapeo,cols_det,err = cruzar_trazabilidad(excel_ref["path"], res_globales)
                    if err: mapeo=None

                titulo = txt_proy.value.strip() or "Escaneo"
                orden  = [ip for ip,_,_ in gateways]
                opts   = {"strings":chk_strings.value,"alarmas":chk_alarmas.value,
                          "diag":chk_diag.value,"voltaje":chk_voltaje.value,"mac":chk_mac.value}
                msgs=[]
                if opcion in("pdf","ambos"):
                    try:
                        r=generar_pdf(res_globales,orden,mapeo,cols_det,
                                      titulo_proyecto=titulo,opts=opts,cbt_por_ip=cbt_por_ip)
                        msgs.append(f"PDF: {os.path.basename(r)}")
                    except Exception as ex: msgs.append(f"Error PDF: {ex}")
                if opcion in("csv","ambos"):
                    try:
                        r=generar_csv(res_globales,orden,mapeo,
                                      titulo_proyecto=titulo,opts=opts,cbt_por_ip=cbt_por_ip)
                        msgs.append(f"CSV: {os.path.basename(r)}")
                    except Exception as ex: msgs.append(f"Error CSV: {ex}")
                lbl_est.value = " | ".join(msgs) if msgs else "Sin exportar"
                page.update()

            threading.Thread(target=run, daemon=True).start()

        def canc(e2): dlg.open=False; page.update()

        grp = ft.RadioGroup(value="ambos", content=ft.Column([
            ft.Radio(value="pdf",   label="Solo PDF"),
            ft.Radio(value="csv",   label="Solo CSV"),
            ft.Radio(value="ambos", label="PDF + CSV"),
        ],spacing=4))
        hay_excel = bool(excel_ref["path"])
        chk_traz  = ft.Checkbox(label="Cruzar con Excel de trazabilidad",
                                value=hay_excel, disabled=not hay_excel, active_color="#38bdf8")
        dlg = ft.AlertDialog(
            modal=True, title=ft.Text("Exportar reporte", weight="bold"),
            content=ft.Column([
                ft.Text("El nombre del archivo será: Título_proyecto + fecha", color="#94a3b8", size=11),
                grp, ft.Divider(color="#334155"), chk_traz,
            ], tight=True, spacing=8),
            actions=[
                ft.TextButton("Cancelar", on_click=canc),
                ft.ElevatedButton("Exportar", on_click=do_exp, bgcolor="#0ea5e9", color="white"),
            ],
            actions_alignment=ft.MainAxisAlignment.END, bgcolor="#1e293b")
        page.dialog=dlg; dlg.open=True; page.update()

    # ══════════════════════════════════════════
    #  LAYOUT
    # ══════════════════════════════════════════
    sec_proy = ft.Container(content=ft.Column([
        ft.Text("PROYECTO / REPORTE", size=11, weight="bold", color="#475569"),
        ft.Row([ft.Icon(ft.icons.FOLDER_OUTLINED,color="#a78bfa",size=18),txt_proy],spacing=8),
        ft.Text("Este nombre aparecerá en la portada del PDF y como nombre del archivo",
                color="#475569",size=11,italic=True),
    ],spacing=6), bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"), border_radius=10, padding=14)

    sec_plantas = ft.Container(content=ft.Column([
        ft.Text("PLANTAS GUARDADAS", size=11, weight="bold", color="#475569"),
        ft.Row([plantas_dropdown,
                ft.ElevatedButton("Cargar", icon=ft.icons.DOWNLOAD, on_click=on_cargar_planta,
                                  bgcolor="#1e293b", color="#38bdf8",
                                  style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155"))),
                ft.ElevatedButton("Guardar actual", icon=ft.icons.SAVE, on_click=on_guardar_planta,
                                  bgcolor="#1e293b", color="#4ade80",
                                  style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155"))),
                ft.ElevatedButton("Eliminar", icon=ft.icons.DELETE_OUTLINE, on_click=on_eliminar_planta,
                                  bgcolor="#1e293b", color="#f87171",
                                  style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155"))),
                ],spacing=8,wrap=True),
        lbl_planta,
        ft.Text("Guarda la configuración de gateways + CBT para no tener que repetirla",
                color="#475569",size=11,italic=True),
    ],spacing=6), bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"), border_radius=10, padding=14)

    sec_gw = ft.Container(content=ft.Column([
        ft.Text("GATEWAYS", size=11, weight="bold", color="#475569"),
        ft.Row([txt_ip, txt_ini, txt_fin,
                ft.ElevatedButton("Añadir", icon=ft.icons.ADD, on_click=add_gw,
                                  bgcolor="#1e293b", color="#38bdf8",
                                  style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155")))],
               spacing=8, wrap=True),
        lista_gw,
    ],spacing=8), bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"), border_radius=10, padding=14)

    sec_cbt = ft.Container(content=ft.Column([
        ft.Text("CBT POR IP GATEWAY", size=11, weight="bold", color="#475569"),
        ft.Text("Cada archivo CBT se asocia a una IP de gateway. La MAC se relaciona por IP + Modbus ID.",
                color="#475569",size=11,italic=True),
        ft.Row([cbt_ip_dropdown,
                ft.ElevatedButton("Cargar CBT", icon=ft.icons.FINGERPRINT,
                                  on_click=lambda e: fp_cbt.pick_files(
                                      dialog_title="Archivo CBT del Gateway",
                                      allowed_extensions=["cbt","CBT","txt"]),
                                  bgcolor="#1e293b", color="#a78bfa",
                                  style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155"))),
                lbl_cbt,
                ],spacing=8,wrap=True),
    ],spacing=6), bgcolor="#0f172a", border=ft.border.all(1,"#2d1b69"), border_radius=10, padding=14)

    sec_excel = ft.Container(content=ft.Column([
        ft.Text("TRAZABILIDAD EXCEL (OPCIONAL)", size=11, weight="bold", color="#475569"),
        ft.Row([
            ft.ElevatedButton("Cargar Excel", icon=ft.icons.TABLE_CHART,
                              on_click=lambda e: fp_excel.pick_files(
                                  dialog_title="Excel de trazabilidad",
                                  allowed_extensions=["xlsx","xls"]),
                              bgcolor="#1e293b", color="#4ade80",
                              style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155"))),
            ft.TextButton("Quitar", on_click=lambda e: setattr(excel_ref,"path",None) or
                          setattr(lbl_excel,"value","Sin Excel") or
                          setattr(lbl_excel,"color","#64748b") or page.update(),
                          style=ft.ButtonStyle(color="#64748b")),
            lbl_excel,
        ],spacing=8,wrap=True),
        ft.Text("Detecta columnas IP, ID y N.Caja automáticamente",color="#475569",size=11,italic=True),
    ],spacing=6), bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"), border_radius=10, padding=14)

    sec_opts = ft.Container(content=ft.Column([
        ft.Text("OPCIONES DE ESCANEO", size=11, weight="bold", color="#475569"),
        ft.Row([chk_strings, txt_umbral], spacing=16, wrap=True),
        ft.Row([chk_alarmas, chk_diag], spacing=16, wrap=True),
        ft.Row([chk_voltaje, chk_mac], spacing=16, wrap=True),
        ft.Text("Las columnas del reporte se generan solo si la opción está activada",
                color="#475569",size=11,italic=True),
    ],spacing=6), bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"), border_radius=10, padding=14)

    sec_res = ft.Container(content=ft.Column([
        ft.Text("RESULTADOS", size=11, weight="bold", color="#475569"),
        prog, lbl_est, seccion_stats,
        ft.Container(content=ft.Column([tabla],scroll=ft.ScrollMode.AUTO),
                     height=360, border=ft.border.all(1,"#1e293b"), border_radius=8),
    ],spacing=8), bgcolor="#0f172a", border=ft.border.all(1,"#1e293b"), border_radius=10, padding=14)

    card = ft.Container(
        content=ft.Column([
            ft.Row([ft.Icon(ft.icons.WIFI_TETHERING,color="#38bdf8",size=26),
                    ft.Text("WEBDOM MODBUS SCANNER",size=21,weight="bold",color="#f1f5f9")],spacing=10),
            sec_proy, sec_plantas, sec_gw, sec_cbt, sec_excel, sec_opts,
            ft.Row([
                ft.ElevatedButton("Escanear", icon=ft.icons.PLAY_ARROW, on_click=escanear,
                                  bgcolor="#0ea5e9", color="white",
                                  style=ft.ButtonStyle(elevation=2)),
                ft.ElevatedButton("Cancelar", icon=ft.icons.STOP, on_click=cancelar,
                                  bgcolor="#1e293b", color="#f87171",
                                  style=ft.ButtonStyle(side=ft.BorderSide(1,"#334155"))),
            ],spacing=10),
            sec_res,
        ],spacing=14),
        width=820, padding=20, bgcolor="#0f172a",
        border=ft.border.all(1,"#1e293b"), border_radius=16,
        shadow=ft.BoxShadow(blur_radius=24, color="#00000066", offset=ft.Offset(0,4)))

    page.add(ft.Column([card], horizontal_alignment=ft.CrossAxisAlignment.CENTER))

ft.app(target=main)